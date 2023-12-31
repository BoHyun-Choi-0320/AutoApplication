import os
import shutil
import sys

import openpyxl.styles
import pandas as pd
import xlwings as xw
from PyPDF2 import PdfMerger, PdfReader
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import *
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor, TwoCellAnchor, AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D, XDRPoint2D
from openpyxl.styles import Font, Border, Side
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU


class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('My Auto Application')
        self.setICon()
        self.setGridLayout()
        self.setWindowCenter()
        self.show()

    def setICon(self):
        self.setWindowIcon(QIcon('./img/Geostory_Logo.png'))
        self.setGeometry(600, 300, 500, 300)
        self.show()

    def setGridLayout(self):
        grid = QGridLayout()
        self.setLayout(grid)

        # 객체 생성
        # self.filePath = QLineEdit('E:/geostory/2023/타 부서 업무협조/해양사업부/데이터/총괄표/(신양식) 장애물 관리대장(무안공항)_231214_test_.xlsx')
        # self.folderPath = QLineEdit('E:/geostory/2023/타 부서 업무협조/해양사업부/데이터/FolderTree_v2')
        # self.formPath = QLineEdit('E:/geostory/2023/타 부서 업무협조/해양사업부/데이터/장애물 관리대장 양식')

        self.filePath = QLineEdit()
        self.folderPath = QLineEdit()
        self.formPath = QLineEdit()

        self.fileSelectBtn = QPushButton('열기')
        self.folderSelectBtn = QPushButton('열기')
        self.formSelectBtn = QPushButton('열기')
        self.accessBtn = QPushButton('적용')
        self.mergedPDFBtn = QPushButton('PDF 병합')

        # 객체 위치 지정
        grid.addWidget(QLabel('관리대장 : '), 1, 1)
        grid.addWidget(QLabel('이미지 경로 : '), 2, 1)
        grid.addWidget(QLabel('관리대장 양식 : '), 3, 1)

        grid.addWidget(self.filePath, 1, 2)
        grid.addWidget(self.folderPath, 2, 2)
        grid.addWidget(self.formPath, 3, 2)

        grid.addWidget(self.fileSelectBtn, 1, 3)
        grid.addWidget(self.folderSelectBtn, 2, 3)
        grid.addWidget(self.formSelectBtn, 3, 3)

        grid.addWidget(self.accessBtn, 4, 2)
        grid.addWidget(self.mergedPDFBtn, 4, 3)

        self.fileSelectBtn.clicked.connect(self.selectFile)
        self.folderSelectBtn.clicked.connect(self.selectFolder)
        self.formSelectBtn.clicked.connect(self.selectForm)
        self.accessBtn.clicked.connect(self.accessLogic)

    def setWindowCenter(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def selectFile(self):
        fName = QFileDialog.getOpenFileName(self, '파일 선택', '', 'excel file(*.xls *.xlsx')
        self.filePath.setText(fName[0])

    def selectFolder(self):
        fName = QFileDialog.getExistingDirectory(self, '폴더 선택', '')
        self.folderPath.setText(fName[:])

    def selectForm(self):
        fName = QFileDialog.getExistingDirectory(self, '폴더 선택', '')
        self.formPath.setText(fName[:])

    # 필요한 파일 및 폴더 입력 후 처리 함수
    def accessLogic(self):
        if not self.filePath.text() or not self.folderPath.text() or not self.formPath.text():
            print('파일 또는 폴더를 선택해 주세요.')
        else:
            print('파일, 폴더 입력 완료', self.filePath.text(), self.formPath.text())
            self.readExcel()

    # 관리대장 읽어오는 함수
    def readExcel(self):
        self.pf = pd.read_excel(self.filePath.text(), header=3,
                                usecols='B,C,E,I,J,K,P,Q,R,S,T,U,V,W,X,Y,Z,AA,AB,AC,AD,AH,AI,AJ,AL,AM,AN,AO,AP,AZ,BA,BO,BR,BS,BW,BY,CA')
        self.pf.columns = ['순번', '신규연도', '연번', '세부종류', '장애물 용도', '명칭', '특례 장애물 구분', '차폐 기준 장애물 및 지정일',
                           '주소1', '주소2',
                           '주소3',
                           '주소4', '도로명주소', '위치구역', '위도', '경도', 'X축', 'Y축', '지반높이', '건물/시설물/수목 높이', '전체높이',
                           '제한표면',
                           '제한표면 침범높이',
                           '협의높이', '위반여부', '건축주', '기관명', '연락처', '관리번호', '건축허가일', '준공승인일', '장애물 등재일', '장애물 등재일2',
                           '장애물 등재일3', '비고', '좌표/높이 결정방법', 'AIP']
        self.makeFolder()
        self.copyExcel()
        try:

            if os.path.isdir(self.pdfPath + "/test"):
                shutil.rmtree(self.pdfPath + "/test")
        except Exception as e:
            print("Error - readExcel : ", e)

    # 결과 저장 폴더 만드는 함수
    def makeFolder(self):
        print('MyApp - makeFolder()')
        self.path = os.getcwd()
        try:
            if not os.path.exists(self.path + "/result"):
                os.mkdir(self.path + "/result")
        except:
            print('Error : Creating directory.' + self.pdfPath)

        self.savePath = self.path + "/result/excel"
        try:
            if not os.path.exists(self.savePath):
                os.mkdir(self.savePath)
        except:
            print('Error : Creating directory.' + self.savePath)

        self.pdfPath = os.path.dirname(self.savePath) + "/pdf"
        try:
            if not os.path.exists(self.pdfPath):
                os.mkdir(self.pdfPath)
        except:
            print('Error : Creating directory.' + self.pdfPath)
        finally:
            self.mergedPDFBtn.clicked.connect(self.mergedPDF)

        try:
            if not os.path.exists(self.pdfPath + "/test"):
                os.mkdir(self.pdfPath + "/test")
        except:
            print('Error : Creating directory.' + self.pdfPath + "/test")

    def copyExcel(self):
        print('MyApp - copyExcel()')
        try:
            for i in range(len(self.pf)):
                imgPath = self.folderPath.text() + "/" + str(self.pf['연번'][i])
                if self.pf['세부종류'][i] == '나무':
                    fileName = '/' + str(self.pf['연번'][i]) + ".xlsx"
                    save = self.savePath + fileName
                    form = self.formPath.text() + "/장애물 관리대장 및 상세표 양식(나무)_v3.xlsx"
                    shutil.copy(form, save)
                    self.inputDataToExcel(savePath=save, data=self.pf.loc[i], imgPath=imgPath)
                elif self.pf['세부종류'][i] == '산':
                    fileName = '/' + str(self.pf['연번'][i]) + ".xlsx"
                    save = self.savePath + fileName
                    form = self.formPath.text() + "/장애물 관리대장 및 상세표 양식(산)_v3.xlsx"
                    shutil.copy(form, save)
                    self.inputDataToExcel(savePath=save, data=self.pf.loc[i], imgPath=imgPath)
                elif self.pf['세부종류'][i] == '건물':
                    fileName = '/' + str(self.pf['연번'][i]) + ".xlsx"
                    save = self.savePath + fileName
                    form = self.formPath.text() + "/장애물 관리대장 및 상세표 양식(건물)_v3.xlsx"
                    shutil.copy(form, save)
                    self.inputDataToExcel(savePath=save, data=self.pf.loc[i], imgPath=imgPath)
                else:
                    fileName = '/' + str(self.pf['연번'][i]) + ".xlsx"
                    save = self.savePath + fileName
                    form = self.formPath.text() + "/장애물 관리대장 및 상세표 양식(기타)_v3.xlsx"
                    shutil.copy(form, save)
                    self.inputDataToExcel(savePath=save, data=self.pf.loc[i], imgPath=imgPath)
        except Exception as e:
            print('Error : Copy Excel. ' + e)

        finally:
            print('MyApp - copyExcel() done!')

    def inputDataToExcel(self, savePath, data, imgPath):
        try:
            print('inputDataToExcel : ', str(data['연번']))
            wb = load_workbook(savePath)
            ws1 = wb["연번"]
            ws2 = wb["연번(장애물 관리대장 상세표)"]
            ws1.title = str(data['연번'])
            ws2.title = str(data['연번']) + "(장애물 관리대장 상세표)"

            ws1['A5'].value = data['연번']
            ws1['B5'].value = data['명칭']
            ws1['C5'].value = data['주소1'].strip() + " " + data['주소2'].strip() + " " + data['주소3'].strip() + " " + (data['주소4'].strip())
            ws1['D6'].value = data['도로명주소']
            ws1['F5'].value = data['위치구역']
            ws1['A9'].value = data['위도']
            ws1['B9'].value = data['경도']
            ws1['C9'].value = data['X축']
            ws1['A9'].number_format = '00"˚"00"′"00.00"″"'
            ws1['D9'].value = data['Y축']
            ws1['B9'].number_format = '00"˚"00"′"00.00"″"'
            ws1['E9'].value = data['좌표/높이 결정방법']
            if type(data['건축허가일']) == str or type(data['건축허가일']) == int or type(data['건축허가일']) == float:
                ws1['F9'].value = str(data['건축허가일'])
            else:
                ws1['F9'].value = (data['건축허가일']).strftime('%y/%m/%d')
            if type(data['준공승인일']) == str or type(data['준공승인일']) == int or type(data['준공승인일']) == float:
                ws1['G9'].value = str(data['준공승인일'])
            else:
                ws1['G9'].value = (data['준공승인일']).strftime('%y/%m/%d')

            ws1['A13'].value = data['지반높이']
            ws1['B13'].value = data['건물/시설물/수목 높이']
            ws1['C13'].value = data['전체높이']
            ws1['D13'].value = data['제한표면']
            ws1['E13'].value = data['제한표면 침범높이']
            if type(data['협의높이']) == int or type(data['협의높이']) == float:
                ws1['F13'].value = data['협의높이']
                ws1['F13'].number_format = '0.00'
            else:
                ws1['F13'].value = str(data['협의높이'])

            ws1['G13'].value = data['위반여부']
            if type(data['신규연도']) == str or type(data['신규연도']) == int or type(data['신규연도']) == float:
                ws1['A18'].value = str(data['신규연도'])
            else:
                ws1['A18'].value = data['신규연도'].strftime('%y/%m/%d')

            ws1['B18'].value = data['특례 장애물 구분']
            ws1['C18'].value = data['차폐 기준 장애물 및 지정일']
            if data['장애물 등재일'] == '-' and data['장애물 등재일2'] =='-' and data['장애물 등재일3']=='-':
                ws1['E18'].value = str(data['장애물 등재일'])
            else:
                ws1['E18'].value = str(data['장애물 등재일']) + " " + str(data['장애물 등재일2']) + " " + str(data['장애물 등재일3'])
            ws1['A23'].value = data['장애물 용도']
            ws1['B23'].value = data['건축주']
            ws1['C23'].value = data['기관명']
            ws1['D23'].value = data['연락처']
            ws1['F23'].value = data['관리번호']
            ws1['A27'].value = data['비고']
            ws1['G5'].value =data['AIP']

            ws1['G20'].border = Border(right=Side(border_style='medium', color="000000"),
                                       bottom=Side(border_style='thin', color="000000"),
                                       left=Side(border_style='thin', color="000000"),
                                       top=Side(border_style='thin', color="000000"))
            ws1['G21'].border = Border(right=Side(border_style='medium', color="000000"),
                                       bottom=Side(border_style='thin', color="000000"),
                                       left=Side(border_style='thin', color="000000"),
                                       top=Side(border_style='thin', color="000000"))
            ws1['G22'].border = Border(right=Side(border_style='medium', color="000000"),
                                       bottom=Side(border_style='thin', color="000000"),
                                       left=Side(border_style='thin', color="000000"),
                                       top=Side(border_style='thin', color="000000"))
            ws1['G23'].border = Border(right=Side(border_style='medium', color="000000"),
                                       bottom=Side(border_style='thin', color="000000"),
                                       left=Side(border_style='thin', color="000000"),
                                       top=Side(border_style='thin', color="000000"))
            ws1['G24'].border = Border(right=Side(border_style='medium', color="000000"),
                                       bottom=Side(border_style='thin', color="000000"),
                                       left=Side(border_style='thin', color="000000"),
                                       top=Side(border_style='thin', color="000000"))
            ws1.font = Font(name='돋움', size=11)
            ws1['A9'].font = Font(name='돋움', size=10)
            ws1['B9'].font = Font(name='돋움', size=10)
            ws1['C9'].font = Font(name='돋움', size=10)
            ws1['D9'].font = Font(name='돋움', size=10)
            ws1['E9'].font = Font(name='돋움', size=10)
            ws1['F9'].font = Font(name='돋움', size=10)
            ws1['G9'].font = Font(name='돋움', size=10)

            ws1.column_dimensions['B'].width = ws1.column_dimensions['A'].width
            if data['순번'] != '제거':
                self.typeImage(data['세부종류'], imgPath, data, wb, ws1, ws2, savePath)

            wb.save(savePath)

            self.excelToPDF(savePath, data['연번'])
        except Exception as e:
            print("inputDataToExcel() - Error! : ", e)
        finally:
            print('inputDataToExcel Done!')

    def typeImage(self, type, imgPath, data, wb, ws1, ws2, savePath):
        print('MyApp - typeImage()')
        try:
            self.setImage(wb=wb, ws=ws1, imgPath=imgPath, data=data, imgType="/현장사진_", savePath=savePath,
                          posRange="A25:G25", position="A25")
            if type == '나무':
                self.setImage(wb=wb, ws=ws2, imgPath=imgPath, data=data, imgType="/단면도_", savePath=savePath,
                              posRange="A5:G16", position="A5")
                self.setImage(wb=wb, ws=ws2, imgPath=imgPath, data=data, imgType="/포인트클라우드_", savePath=savePath,
                              posRange="A20:C20", position="A20")
                self.setImage(wb=wb, ws=ws2, imgPath=imgPath, data=data, imgType="/지상라이다_", savePath=savePath,
                              posRange="E20:G20", position="E20")
            elif type == '산':
                self.setImage(wb=wb, ws=ws2, imgPath=imgPath, data=data, imgType="/단면도_", savePath=savePath,
                              posRange="A5:G16", position="A5")
                self.setImage(wb=wb, ws=ws2, imgPath=imgPath, data=data, imgType="/포인트클라우드_", savePath=savePath,
                              posRange="A20:C20", position="A20")
                self.setImage(wb=wb, ws=ws2, imgPath=imgPath, data=data, imgType="/수치표고자료_", savePath=savePath,
                              posRange="E20:G20", position="E20")
            elif type == '건물':
                self.setImage(wb=wb, ws=ws2, imgPath=imgPath, data=data, imgType="/정사영상_", savePath=savePath,
                              posRange="A5:C16", position="A5")
                self.setImage(wb=wb, ws=ws2, imgPath=imgPath, data=data, imgType="/3D모델링_", savePath=savePath,
                              posRange="E5:G16", position="E5")
                self.setImage(wb=wb, ws=ws2, imgPath=imgPath, data=data, imgType="/단면도_", savePath=savePath,
                              posRange="A20:G20", position="A20")
            elif type == '기타' or type == '철탑':
                self.setImage(wb=wb, ws=ws2, imgPath=imgPath, data=data, imgType="/위치도_", savePath=savePath,
                              posRange="A5:G17", position="A5")
                self.setImage(wb=wb, ws=ws2, imgPath=imgPath, data=data, imgType="/단면도_", savePath=savePath,
                              posRange="A20:G20", position="A20")
        except Exception as e:
            print("typeImage - Error : ", e)
        finally:
            print('typeImage Done!')

    def setImage(self, wb, ws, imgPath, data, imgType, savePath, posRange, position):
        try:
            total_width, total_height = self.getMergedWidthHegiht(posRange, ws)

            imgPath = imgPath + imgType + str(data['연번']) + ".jpg"
            img = Image(imgPath)
            img.width, img.height = self.get_col_width_row_height(total_width, total_height)

            origin_cell = ws[position]

            p2e = pixels_to_EMU
            c2e = cm_to_EMU

            cellh = lambda x: c2e((x * 49.77) / 99)
            cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)

            size = XDRPositiveSize2D(p2e(img.width), p2e(img.height))

            column = origin_cell.column - 1
            coloffset = cellw(0.09)
            row = origin_cell.row - 1
            rowoffset = cellh(0.5)

            marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
            img.anchor = OneCellAnchor(_from=marker, ext=size)

            ws.add_image(img)

            wb.save(savePath)
        except Exception as e:
            print("setImage - Error : ", imgType, " 이미지가 없습니다.", e)
        finally:
            print("setImage Done!")

    def getMergedWidthHegiht(self, posRange, ws):
        try:
            merged_cell_address = posRange
            merged_cell = ws[merged_cell_address]
            start_column, start_row, end_column, end_row = merged_cell[0][0].column, merged_cell[0][0].row, \
                                                           merged_cell[-1][
                                                               -1].column, merged_cell[-1][-1].row
            total_width, total_height = 0, 0
            for col in range(start_column, end_column + 1):
                cell = ws.cell(row=start_row, column=col)
                start_cell_column_letter = openpyxl.utils.get_column_letter(col)
                col_width = ws.column_dimensions[start_cell_column_letter].width
                total_width += col_width * 0.21

            for row in range(start_row, end_row + 1):
                cell = ws.cell(row=row, column=start_column)
                total_height += ws.row_dimensions[row].height * 0.035

            return (total_width, total_height)
        except Exception as e:
            print('Error getMergedWidthHegiht() : ', e)

    def get_col_width_row_height(self, img_width, img_height):
        col_width = (img_width * 7300) / 193 - 5
        row_height = (img_height * 7300) / 193 - 10
        return (col_width, row_height)

    def excelToPDF(self, excelPath, filename):
        print('MyApp - excelToPDF()')
        try:
            app = xw.App(visible=False)
            book = xw.Book(excelPath)
            pdf1 = self.pdfPath + "/test/" + str(filename) + "_1.pdf"
            pdf2 = self.pdfPath + "/test/" + str(filename) + "_2.pdf"
            book.sheets[0].api.ExportAsFixedFormat(0, pdf1)
            book.sheets[1].api.ExportAsFixedFormat(0, pdf2)
            app.kill()

            pdfFile1 = open(pdf1, "rb")
            pdfFile2 = open(pdf2, "rb")

            pdfMerger = PdfMerger()

            pdfMerger.append(PdfReader(pdfFile1))
            pdfMerger.append(PdfReader(pdfFile2))
            pdfFile1.close()
            pdfFile2.close()

            pdfMerger.write(self.pdfPath + "/" + str(filename) + ".pdf")
            pdfMerger.close()
        except Exception as e:
            print('excelToPDF Error : ' + e)
        finally:
            print('excelToPDF Done!')

    def mergedPDF(self):
        print('mergedPDF')
        try:
            if self.pdfPath == None:
                print("값 입력하세요.")
            else:
                pdfList = os.listdir(self.pdfPath)
                sorted(pdfList, key=lambda x: int(x.split('.')[0]))
                pdfList = [file for file in sorted(pdfList, key=lambda x: int(x.split('.')[0])) if
                           file.endswith(".pdf")]

                merger = PdfMerger()

                for pdf in pdfList:
                    merger.append(self.pdfPath + "/" + pdf)

                merger.write(self.pdfPath + "/" + "관리대장.pdf")
                merger.close()

        except Exception as e:
            print("Error - ", e)



if __name__ == '__main__':
    # Application 객체 생성
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())
